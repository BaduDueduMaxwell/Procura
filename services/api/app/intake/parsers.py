import csv
import io
import re
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.config import Settings
from app.domain.errors import FileValidationError
from app.domain.models import IntakeLine
from app.intake.interfaces import ParsedIntake

HEADER_ALIASES = {
    "medicine": "medicine_name", "medicine name": "medicine_name", "drug": "medicine_name",
    "product": "medicine_name", "item": "medicine_name", "generic name": "medicine_name",
    "molecule": "medicine_name", "active ingredient": "medicine_name", "brand": "brand_name",
    "brand name": "brand_name", "strength": "strength", "concentration": "strength",
    "dosage form": "dosage_form", "formulation": "dosage_form", "form": "dosage_form",
    "quantity": "quantity", "requested quantity": "quantity", "qty": "quantity",
    "unit": "unit", "units": "unit", "pack size": "pack_size", "pack": "pack_size",
    "destination": "destination", "market": "destination", "country": "destination",
    "delivery days": "max_lead_time_days", "lead time": "max_lead_time_days",
    "lead time days": "max_lead_time_days", "maximum lead time": "max_lead_time_days",
    "currency": "currency", "cold chain": "cold_chain_required",
}
ALLOWED_MIME = {
    ".csv": {"text/csv", "text/plain", "application/csv", "application/octet-stream", None, ""},
    ".xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/octet-stream", None, ""},
}


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def safe_filename(value: str) -> str:
    name = Path(value).name
    cleaned = re.sub(r"[^A-Za-z0-9._ -]", "_", name).strip(". ")
    if not cleaned or len(cleaned) > 180:
        raise FileValidationError("Use a filename shorter than 180 characters")
    return cleaned


def _number(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        number = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError) as exc:
        raise FileValidationError(f"Expected a whole number, received {value!s}") from exc
    if number <= 0 or not number.is_integer():
        raise FileValidationError(f"Expected a positive whole number, received {value!s}")
    return int(number)


def _boolean(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "required", "cold"}


def _clean(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = " ".join(str(value).split())
    return text[:500] or None


def _line(row_number: int, sheet_name: str | None, headers: list[str], values: list[Any]) -> IntakeLine:
    mapped: dict[str, Any] = {}
    original: dict[str, str | int | float | bool | None] = {}
    for index, header in enumerate(headers):
        if not header or index >= len(values):
            continue
        value = values[index]
        original[header] = value if isinstance(value, (str, int, float, bool)) or value is None else str(value)
        field = HEADER_ALIASES.get(header)
        if field and field not in mapped:
            mapped[field] = value
    for field in ("quantity", "pack_size", "max_lead_time_days"):
        mapped[field] = _number(mapped.get(field))
    for field in ("medicine_name", "brand_name", "strength", "dosage_form", "unit", "destination", "currency"):
        mapped[field] = _clean(mapped.get(field))
    mapped["cold_chain_required"] = _boolean(mapped.get("cold_chain_required"))
    return IntakeLine(source_row=row_number, sheet_name=sheet_name, original_values=original, **mapped)


class SpreadsheetParser:
    def __init__(self, settings: Settings):
        self.max_bytes = settings.intake_max_file_bytes
        self.max_rows = settings.intake_max_rows

    def parse(self, filename: str, content: bytes, content_type: str | None = None) -> ParsedIntake:
        filename = safe_filename(filename)
        suffix = Path(filename).suffix.lower()
        if suffix not in ALLOWED_MIME:
            raise FileValidationError("Upload an .xlsx or .csv procurement list")
        if content_type not in ALLOWED_MIME[suffix]:
            raise FileValidationError("The file content type does not match its extension")
        if not content:
            raise FileValidationError("The uploaded file is empty")
        if len(content) > self.max_bytes:
            raise FileValidationError(f"The file exceeds the {self.max_bytes // (1024 * 1024)} MB limit")
        if suffix == ".xlsx":
            return self._xlsx(filename, content)
        return self._csv(filename, content)

    def _csv(self, filename: str, content: bytes) -> ParsedIntake:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise FileValidationError("Save the CSV using UTF-8 encoding") from exc
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(io.StringIO(text), dialect))
        lines = self._rows_to_lines(rows, None)
        return ParsedIntake(source_type="csv", filename=filename, lines=lines, sheet_names=[])

    def _xlsx(self, filename: str, content: bytes) -> ParsedIntake:
        if not content.startswith(b"PK"):
            raise FileValidationError("The file is not a valid XLSX workbook")
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                infos = archive.infolist()
                if len(infos) > 500 or sum(item.file_size for item in infos) > 25 * 1024 * 1024:
                    raise FileValidationError("The workbook expands beyond the safe processing limit")
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
        except (zipfile.BadZipFile, KeyError, ValueError) as exc:
            raise FileValidationError("The workbook is corrupted or unsupported") from exc
        lines: list[IntakeLine] = []
        names: list[str] = []
        for worksheet in workbook.worksheets:
            rows = [[cell.value for cell in row] for row in worksheet.iter_rows()]
            if not any(any(value not in (None, "") for value in row) for row in rows):
                continue
            names.append(worksheet.title)
            lines.extend(self._rows_to_lines(rows, worksheet.title))
            if len(lines) > self.max_rows:
                raise FileValidationError(f"The workbook exceeds the {self.max_rows:,} row limit across all worksheets")
        workbook.close()
        if not lines:
            raise FileValidationError("The workbook contains no procurement rows")
        return ParsedIntake(source_type="xlsx", filename=filename, lines=lines, sheet_names=names)

    def _rows_to_lines(self, rows: list[list[Any]], sheet_name: str | None) -> list[IntakeLine]:
        nonempty = [(index, row) for index, row in enumerate(rows, start=1) if any(value not in (None, "") for value in row)]
        if len(nonempty) < 2:
            raise FileValidationError("Include a header row and at least one procurement row")
        header_number, header_row = nonempty[0]
        headers = [normalize_header(value) for value in header_row]
        if not any(header in HEADER_ALIASES for header in headers):
            raise FileValidationError("No recognized procurement columns were found")
        data_rows = nonempty[1:]
        if len(data_rows) > self.max_rows:
            raise FileValidationError(f"The file exceeds the {self.max_rows:,} row limit")
        lines: list[IntakeLine] = []
        for row_number, row in data_rows:
            for value in row:
                if isinstance(value, str) and value.lstrip().startswith("="):
                    raise FileValidationError(f"Formula found on row {row_number}; replace formulas with values")
            lines.append(_line(row_number, sheet_name, headers, row))
        if header_number > 20:
            raise FileValidationError("Move the header row into the first 20 rows")
        return lines
